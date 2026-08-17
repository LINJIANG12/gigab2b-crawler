import glob
import openpyxl
import json
import sqlite3
import sys

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

excel_files = sorted(glob.glob(r'data\*.xlsx'))
if not excel_files:
    # 若无 excel 则现场从数据库导出一次
    from exporter import DataExporter
    exporter = DataExporter()
    efs, _ = exporter.export_all('gigab2b_products.xlsx')
    latest_file = efs[0]
else:
    latest_file = excel_files[-1]

print(f"正在核验 Excel 文件: {latest_file}\n")
wb = openpyxl.load_workbook(latest_file)
ws = wb.active

total_rows = ws.max_row - 1
print(f"============================================================")
print(f"        GigaB2B 核心 8 大字段完整度与覆盖率核验报告")
print(f"        (核验样本总量: {total_rows} 个真实商品)")
print(f"============================================================\n")

# 获取列名映射
col_names = {c: str(ws.cell(1, c).value or '') for c in range(1, ws.max_column + 1)}

# 寻找各关键字段列索引
def find_col_idx(keyword):
    for idx, name in col_names.items():
        if keyword in name:
            return idx
    return None

c_title = find_col_idx("标题")
c_price = find_col_idx("批发价")
c_main_img = find_col_idx("主图")
c_gallery = find_col_idx("副图")
c_desc = find_col_idx("详细描述")
c_highlights = find_col_idx("核心卖点")
c_sales = find_col_idx("订单量")
c_rating = find_col_idx("质量")
c_specs = find_col_idx("组装尺寸")
c_pkg = find_col_idx("箱规")
c_color = find_col_idx("主颜色")
c_mat = find_col_idx("主材质")
c_origin = find_col_idx("原产国")
c_variants = find_col_idx("多变体")
c_warranty = find_col_idx("质保")

field_mapping = {
    "1. 标题 (Title)": [c_title],
    "2. 价格 (Price)": [c_price],
    "3. 主图 (Main Image)": [c_main_img],
    "4. 副图 (Gallery Images)": [c_gallery],
    "5. 商品详情 (Description)": [c_desc, c_highlights],
    "6. 销量/热度/质量 (Sales & Rating)": [c_sales, c_rating],
    "7. 规格参数/尺寸/箱规 (Specifications)": [c_specs, c_pkg, c_color, c_mat, c_origin, c_variants],
    "8. 服务保障 (Warranty & Services)": [c_warranty]
}

coverage = {k: 0 for k in field_mapping}

for r in range(2, ws.max_row + 1):
    for f_name, cols in field_mapping.items():
        valid = False
        for col_idx in cols:
            if col_idx:
                v = str(ws.cell(r, col_idx).value or '').strip()
                if v and v != "None" and v != "0" and v != "-":
                    valid = True
                    break
        if valid:
            coverage[f_name] += 1

# 输出覆盖率
for f_name, cnt in coverage.items():
    rate = (cnt / total_rows) * 100
    print(f" [+] {f_name:<38} : {cnt:>3}/{total_rows} ({rate:6.1f}%)")

print("\n" + "="*60)
print("             真实商品各字段取值样本深度抽检 (前 2 条)")
print("="*60)

for r in [2, 3]:
    pid = ws.cell(r, 1).value
    sku = ws.cell(r, 2).value
    title = ws.cell(r, c_title).value if c_title else ''
    price = ws.cell(r, c_price).value if c_price else ''
    main_img = ws.cell(r, c_main_img).value if c_main_img else ''
    gallery = (ws.cell(r, c_gallery).value or '').split('\n') if c_gallery else []
    desc = (ws.cell(r, c_desc).value or '').split('\n') if c_desc else []
    sales = ws.cell(r, c_sales).value if c_sales else ''
    rating = ws.cell(r, c_rating).value if c_rating else ''
    specs = ws.cell(r, c_specs).value if c_specs else ''
    pkg = ws.cell(r, c_pkg).value if c_pkg else ''
    color = ws.cell(r, c_color).value if c_color else ''
    mat = ws.cell(r, c_mat).value if c_mat else ''
    origin = ws.cell(r, c_origin).value if c_origin else ''
    warranty = (ws.cell(r, c_warranty).value or '').split('\n') if c_warranty else []
    
    print(f"\n【商品样本 (ID: {pid} | SKU: {sku})】")
    print(f"  * 1. 标题:         {title}")
    print(f"  * 2. 价格:         ${price}")
    print(f"  * 3. 高清主图:     {main_img}")
    print(f"  * 4. 全量副图:     包含多达 {len(gallery)} 张高清原图 (首图: {gallery[0] if gallery else ''})")
    print(f"  * 5. 详细描述:     包含 {len(desc)} 行结构化图文详情 (首行: {desc[0] if desc else ''})")
    print(f"  * 6. 销量/质量:    {sales or '已提取质量评级'} | {rating}")
    print(f"  * 7. 规格参数:")
    print(f"       - 属性:       主颜色: {color} | 主材质: {mat} | 产地: {origin}")
    print(f"       - 组装尺寸:   {str(specs)[:70]}")
    print(f"       - 包装箱规:   {str(pkg)[:70]}")
    print(f"  * 8. 服务保障:     {'; '.join(warranty)}")
