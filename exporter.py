import os
import csv
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
from config import DATA_DIR, IMAGE_DIR
from database import Database

class DataExporter:
    """
    全量数据导出模块（全量 37 字段独立列完整版）：
    涵盖基础属性、价格体系、分仓库存、代发运费、运输时效、LTL物流、体积毛重、
    规格箱规、文档手册、质保政策、全量副图、多变体与在售状态说明。
    """
    def __init__(self, output_dir: str = DATA_DIR):
        self.output_dir = output_dir
        self.db = Database.get_instance()
        os.makedirs(self.output_dir, exist_ok=True)

    def _safe_save_workbook(self, wb: openpyxl.Workbook, target_path: str) -> str:
        """带自动避让锁定的安全保存函数"""
        try:
            wb.save(target_path)
            return target_path
        except PermissionError:
            name_p, ext_p = os.path.splitext(target_path)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_path = f"{name_p}_{ts}{ext_p}"
            wb.save(safe_path)
            return safe_path

    def export_all(self, excel_filename: str = "gigab2b_products.xlsx", csv_filename: str = "gigab2b_products.csv") -> tuple[list[str], str]:
        """从数据库全量导出所有已采集商品"""
        items = self.db.get_all_products()
        if not items:
            return [], ""

        excel_files = self.export_to_excel_chunked(items, excel_filename)
        csv_file = self.export_to_csv(items, csv_filename)
        return excel_files, csv_file

    def export_to_excel_chunked(self, items: list[dict], base_filename: str = "gigab2b_products.xlsx", chunk_size: int = 50000) -> list[str]:
        """分卷导出结构化 Excel 表格"""
        if not items:
            return []

        # 37 个全量独立列
        columns = [
            ("product_id", "商品ID (Product ID)"),
            ("sku", "货号 / SKU"),
            ("title", "商品标题 (Title)"),
            ("category_path", "分类全路径 (Category)"),
            ("store_name", "店铺/卖家名称 (Store Name)"),
            ("store_code", "店铺代码 (Store Code)"),
            ("product_status", "在售与权限状态 (Product Status)"),
            ("status_reason", "数据状态与原因说明 (Status Reason)"),
            ("price", "B2B批发价 (Price)"),
            ("discount_price", "折扣优惠价 (Discount Price)"),
            ("original_price", "市场参考价 (MSRP)"),
            ("moq", "起订量 (MOQ)"),
            ("total_stock", "总现货库存 (Total Stock)"),
            ("inventory_warehouses", "各海外分仓库存 (Warehouses Stock)"),
            ("drop_ship_fee", "一件代发预估运费 (Drop Ship Fee)"),
            ("cloud_freight_range", "云仓物流运费区间 (Cloud Freight)"),
            ("handling_time", "出库处理时效 (Handling Time)"),
            ("delivery_time", "运输派送时效 (Transit Time)"),
            ("is_ltl", "是否LTL大件物流 (Is LTL)"),
            ("total_weight", "商品总磅重 (Weight)"),
            ("total_volume", "商品总体积 (Volume ft³)"),
            ("main_color", "主颜色 (Main Color)"),
            ("main_material", "主材质 (Main Material)"),
            ("origin_place", "原产国 (Origin Place)"),
            ("upc", "UPC / 条形码"),
            ("product_dimensions", "产品组装尺寸 (Dimensions)"),
            ("package_size", "外包装箱规/毛重 (Cartons & Weight)"),
            ("rating", "平台质量/退货率 (Quality Rating)"),
            ("sales_count", "采购/订单量 (Orders/Sales)"),
            ("shipping_and_services", "服务与质保政策 (Warranty & Service)"),
            ("documents", "说明书/指导文档 (Manuals)"),
            ("bullet_points", "核心卖点提要 (Highlights)"),
            ("description_text", "商品详细描述 (Description)"),
            ("main_image", "高清主图链接 (Main Image)"),
            ("gallery_images", "全量副图链接 (Gallery Images)"),
            ("variants", "多变体属性明细 (Variants)"),
            ("url", "商品直达链接 (URL)")
        ]

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        exported_files = []
        total_items = len(items)
        num_chunks = (total_items + chunk_size - 1) // chunk_size

        for chunk_idx in range(num_chunks):
            start_i = chunk_idx * chunk_size
            end_i = min(start_i + chunk_size, total_items)
            chunk_data = items[start_i:end_i]

            if num_chunks == 1:
                cur_filename = base_filename
            else:
                name_part, ext = os.path.splitext(base_filename)
                cur_filename = f"{name_part}_part{chunk_idx + 1}{ext}"

            filepath = os.path.join(self.output_dir, cur_filename)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "商品全量数据"

            # 写入表头
            headers = [col[1] for col in columns]
            ws.append(headers)

            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 28

            # 写入数据行
            for row_idx, item in enumerate(chunk_data, start=2):
                row_data = []
                for field, _ in columns:
                    val = item.get(field, "")
                    if isinstance(val, list):
                        if field == "gallery_images":
                            val_str = "\n".join(val)
                        elif field == "bullet_points":
                            val_str = "\n".join([f"• {b}" for b in val])
                        elif field == "variants":
                            formatted_vars = []
                            for v in val:
                                if isinstance(v, dict):
                                    title = v.get('title') or v.get('name') or ''
                                    pr = f" (Price: ${v.get('price')})" if v.get('price') else ""
                                    st = f" [{v.get('stock')}]" if v.get('stock') is not None else ""
                                    formatted_vars.append(f"• {title}{pr}{st}")
                                else:
                                    formatted_vars.append(f"• {v}")
                            val_str = "\n".join(formatted_vars)
                        else:
                            val_str = "\n".join([str(x) for x in val])
                    elif isinstance(val, dict):
                        val_str = "\n".join([f"{k}: {v}" for k, v in val.items()])
                    else:
                        val_str = str(val) if val is not None else ""
                    row_data.append(val_str)

                ws.append(row_data)
                ws.row_dimensions[row_idx].height = 22

                for col_idx in range(1, len(row_data) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = Font(name="Microsoft YaHei", size=10)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            # 列宽设置
            for col_idx in range(1, len(columns) + 1):
                col_letter = get_column_letter(col_idx)
                field_name = columns[col_idx - 1][0]
                if field_name in ["title", "description_text", "gallery_images"]:
                    ws.column_dimensions[col_letter].width = 38
                elif field_name in ["status_reason", "product_dimensions", "package_size", "category_path", "inventory_warehouses"]:
                    ws.column_dimensions[col_letter].width = 28
                elif field_name in ["product_status", "main_image", "url", "variants", "shipping_and_services", "documents"]:
                    ws.column_dimensions[col_letter].width = 24
                elif field_name in ["sku", "store_name", "rating", "main_material", "drop_ship_fee", "cloud_freight_range"]:
                    ws.column_dimensions[col_letter].width = 18
                else:
                    ws.column_dimensions[col_letter].width = 14

            actual_saved_path = self._safe_save_workbook(wb, filepath)
            exported_files.append(actual_saved_path)

        return exported_files

    def export_to_csv(self, items: list[dict], filename: str = "gigab2b_products.csv") -> str:
        """导出全字段 UTF-8 BOM CSV 文件"""
        filepath = os.path.join(self.output_dir, filename)
        fieldnames = [
            "product_id", "sku", "title", "category_path", "store_name", "store_code",
            "product_status", "status_reason",
            "price", "discount_price", "original_price", "moq", "currency",
            "total_stock", "inventory_warehouses",
            "drop_ship_fee", "cloud_freight_range", "handling_time", "delivery_time",
            "is_ltl", "total_weight", "total_volume",
            "main_color", "main_material", "origin_place", "upc",
            "product_dimensions", "package_size",
            "rating", "sales_count", "reviews_count",
            "shipping_and_services", "documents",
            "bullet_points", "description_text",
            "main_image", "gallery_images", "variants", "url"
        ]

        try:
            f = open(filepath, "w", encoding="utf-8-sig", newline="")
        except PermissionError:
            name_p, ext_p = os.path.splitext(filepath)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"{name_p}_{ts}{ext_p}"
            f = open(filepath, "w", encoding="utf-8-sig", newline="")

        with f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            for item in items:
                row = {}
                for k in fieldnames:
                    v = item.get(k, "")
                    if isinstance(v, list):
                        row[k] = " | ".join([str(x) for x in v])
                    elif isinstance(v, dict):
                        row[k] = "; ".join([f"{k_}:{v_}" for k_, v_ in v.items()])
                    else:
                        row[k] = str(v) if v is not None else ""
                writer.writerow(row)

        return filepath

    def download_product_images(self, item: dict, session: requests.Session = None) -> list[str]:
        """下载单个商品图片"""
        pid = item.get("product_id") or item.get("sku") or "unknown"
        item_img_dir = os.path.join(IMAGE_DIR, str(pid))
        os.makedirs(item_img_dir, exist_ok=True)

        urls = []
        if item.get("main_image"):
            urls.append(item["main_image"])
        for g_url in item.get("gallery_images", []):
            if g_url not in urls:
                urls.append(g_url)

        saved_files = []
        for idx, img_url in enumerate(urls):
            try:
                ext = img_url.split(".")[-1].split("?")[0]
                if ext.lower() not in ["jpg", "jpeg", "png", "webp", "gif"]:
                    ext = "jpg"
                filename = f"main.{ext}" if idx == 0 else f"gallery_{idx}.{ext}"
                filepath = os.path.join(item_img_dir, filename)

                if os.path.exists(filepath) and os.path.getsize(filepath) > 0:
                    saved_files.append(filepath)
                    continue

                getter = session.get if session else requests.get
                resp = getter(img_url, timeout=15)
                if resp.status_code == 200:
                    with open(filepath, "wb") as f:
                        f.write(resp.content)
                    saved_files.append(filepath)
            except Exception:
                pass

        return saved_files
